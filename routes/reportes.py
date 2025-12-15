from flask import Blueprint, render_template, send_file, request, session, redirect, url_for, flash
import pandas as pd
from fpdf import FPDF
from io import BytesIO
import mysql.connector
from datetime import datetime
import os

# =====================
# CONEXIÓN DIRECTA
# =====================
def obtener_conexion():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="parrilla51"
    )

# =====================
# VERIFICACIÓN DE ADMINISTRADOR
# =====================
def verificar_admin():
    """Función helper para verificar si el usuario es administrador"""
    if 'logueado' not in session:
        return False, 'Debes iniciar sesión primero'
    
    if session.get('rol') != 'administrador':
        return False, f'Acceso denegado. Solo administradores pueden acceder.'
    
    return True, None

# =====================
# BLUEPRINT
# =====================
reportes_bp = Blueprint("reportes", __name__)

# =====================
# REPORTES DE VENTAS
# =====================
@reportes_bp.route("/ventas", methods=["GET", "POST"])
def reportes_ventas():
    es_admin, mensaje = verificar_admin()
    if not es_admin:
        flash(mensaje, 'danger')
        return redirect(url_for('auth.login'))
    
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)

    # Valores por defecto
    busqueda = request.form.get("busqueda", "").strip() if request.method == "POST" else ""
    filtro_mes = request.form.get("mes", "").strip() if request.method == "POST" else ""
    filtro_estado = request.form.get("estado", "").strip() if request.method == "POST" else ""

    # Construir la consulta base
    query = """
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora, p.total, 
               p.estado, p.metodo_pago, p.tipo_entrega
        FROM pedidos p
        INNER JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE 1=1
    """
    params = []

    # Aplicar filtros
    if busqueda:
        query += " AND (u.nombre LIKE %s OR u.apellido LIKE %s OR CAST(p.id_pedido AS CHAR) LIKE %s)"
        busqueda_param = f"%{busqueda}%"
        params.extend([busqueda_param, busqueda_param, busqueda_param])

    if filtro_mes:
        query += " AND DATE_FORMAT(p.fecha, '%%Y-%%m') = %s"
        params.append(filtro_mes)

    if filtro_estado:
        query += " AND p.estado = %s"
        params.append(filtro_estado)

    query += " ORDER BY p.fecha DESC, p.hora DESC"

    cursor.execute(query, tuple(params))
    pedidos = cursor.fetchall()
    
    # Calcular estadísticas
    total_ventas = sum(p['total'] for p in pedidos if p['total'])
    total_pedidos = len(pedidos)
    
    cursor.close()
    conexion.close()

    return render_template(
        "reportes_ventas.html",
        pedidos=pedidos,
        busqueda=busqueda,
        filtro_mes=filtro_mes,
        filtro_estado=filtro_estado,
        total_ventas=total_ventas,
        total_pedidos=total_pedidos
    )

# =====================
# REPORTES DE INVENTARIO
# =====================
@reportes_bp.route("/inventario", methods=["GET", "POST"])
def reportes_inventario():
    es_admin, mensaje = verificar_admin()
    if not es_admin:
        flash(mensaje, 'danger')
        return redirect(url_for('auth.login'))
    
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)

    # Valores por defecto
    busqueda = request.form.get("busqueda", "").strip() if request.method == "POST" else ""
    filtro_categoria = request.form.get("categoria", "").strip() if request.method == "POST" else ""
    filtro_stock = request.form.get("stock", "").strip() if request.method == "POST" else ""

    # Construir la consulta base
    query = """
        SELECT p.id_producto, p.nombre, p.cantidad, p.precio, p.descripcion,
               c.nombre_categoria, c.id_categoria, p.fecha_vencimiento, p.fecha_lote
        FROM productos p
        LEFT JOIN categorias c ON p.cod_categoria = c.id_categoria
        WHERE 1=1
    """
    params = []

    # Aplicar filtros
    if busqueda:
        query += " AND (p.nombre LIKE %s OR p.descripcion LIKE %s)"
        busqueda_param = f"%{busqueda}%"
        params.extend([busqueda_param, busqueda_param])

    if filtro_categoria:
        query += " AND c.id_categoria = %s"
        params.append(int(filtro_categoria))

    if filtro_stock == "bajo":
        query += " AND p.cantidad < 5 AND p.cantidad > 0"
    elif filtro_stock == "sin_stock":
        query += " AND p.cantidad = 0"
    elif filtro_stock == "disponible":
        query += " AND p.cantidad >= 5"

    query += " ORDER BY p.cantidad ASC"

    cursor.execute(query, tuple(params))
    productos = cursor.fetchall()
    
    # Obtener categorías para el filtro
    cursor.execute("SELECT id_categoria, nombre_categoria FROM categorias ORDER BY nombre_categoria")
    categorias = cursor.fetchall()
    
    # Calcular estadísticas
    total_productos = len(productos)
    productos_bajo_stock = sum(1 for p in productos if p['cantidad'] and 0 < p['cantidad'] < 5)
    productos_sin_stock = sum(1 for p in productos if p['cantidad'] == 0)
    valor_inventario = sum(p['cantidad'] * p['precio'] for p in productos if p['cantidad'] and p['precio'])
    
    cursor.close()
    conexion.close()

    return render_template(
        "reportes_inventario.html",
        productos=productos,
        categorias=categorias,
        busqueda=busqueda,
        filtro_categoria=filtro_categoria,
        filtro_stock=filtro_stock,
        total_productos=total_productos,
        productos_bajo_stock=productos_bajo_stock,
        productos_sin_stock=productos_sin_stock,
        valor_inventario=valor_inventario
    )

# =====================
# EXPORTAR VENTAS A EXCEL
# =====================
@reportes_bp.route("/ventas/exportar_excel")
def exportar_ventas_excel():
    es_admin, mensaje = verificar_admin()
    if not es_admin:
        flash(mensaje, 'danger')
        return redirect(url_for('auth.login'))
    
    conexion = obtener_conexion()
    query = """
        SELECT p.id_pedido AS 'ID Pedido', 
               CONCAT(u.nombre, ' ', u.apellido) AS 'Cliente',
               p.fecha AS 'Fecha', 
               p.hora AS 'Hora',
               p.total AS 'Total', 
               p.estado AS 'Estado',
               p.metodo_pago AS 'Método de Pago',
               p.tipo_entrega AS 'Tipo de Entrega'
        FROM pedidos p
        INNER JOIN usuarios u ON p.cod_usuario = u.id_usuario
        ORDER BY p.fecha DESC
    """
    df = pd.read_sql(query, conexion)
    conexion.close()

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ventas")
        
        # Mejorar formato
        workbook = writer.book
        worksheet = writer.sheets["Ventas"]
        
        # Formato de encabezados
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar estilos a encabezados
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
        # Aplicar bordes a todas las celdas
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formato de moneda para columna Total
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=5)  # Columna Total
            cell.number_format = '$#,##0'
    
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"reporte_ventas_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================
# EXPORTAR INVENTARIO A EXCEL
# =====================
@reportes_bp.route("/inventario/exportar_excel")
def exportar_inventario_excel():
    es_admin, mensaje = verificar_admin()
    if not es_admin:
        flash(mensaje, 'danger')
        return redirect(url_for('auth.login'))
    
    conexion = obtener_conexion()
    query = """
        SELECT p.id_producto AS 'ID',
               p.nombre AS 'Producto',
               p.cantidad AS 'Cantidad',
               p.precio AS 'Precio',
               c.nombre_categoria AS 'Categoría',
               p.descripcion AS 'Descripción',
               CASE 
                   WHEN p.cantidad = 0 THEN 'SIN STOCK'
                   WHEN p.cantidad < 5 THEN 'STOCK BAJO'
                   ELSE 'DISPONIBLE'
               END AS 'Estado Stock'
        FROM productos p
        LEFT JOIN categorias c ON p.cod_categoria = c.id_categoria
        ORDER BY p.cantidad ASC
    """
    df = pd.read_sql(query, conexion)
    conexion.close()

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Inventario")
        
        # Mejorar formato
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        workbook = writer.book
        worksheet = writer.sheets["Inventario"]
        
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar estilos a encabezados
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Aplicar colores según estado de stock
        for row in range(2, len(df) + 2):
            estado_cell = worksheet.cell(row=row, column=7)  # Columna Estado Stock
            
            if estado_cell.value == 'SIN STOCK':
                fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                font = Font(bold=True, color="FFFFFF")
            elif estado_cell.value == 'STOCK BAJO':
                fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
                font = Font(bold=True, color="000000")
            else:
                fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                font = Font(bold=True, color="FFFFFF")
            
            estado_cell.fill = fill
            estado_cell.font = font
            
            # Aplicar bordes a todas las celdas
            for cell in worksheet[row]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formato de moneda para columna Precio
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=4)  # Columna Precio
            cell.number_format = '$#,##0'
    
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"reporte_inventario_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================
# PDF PERSONALIZADO
# =====================
class PDF_Parrilla(FPDF):
    def __init__(self, titulo):
        super().__init__()
        self.titulo = titulo
    
    def header(self):
        # Logo (si existe)
        logo_path = os.path.join('static', 'img', 'logooo.png')
        if os.path.exists(logo_path):
            self.image(logo_path, 10, 8, 25)
        
        # Título
        self.set_font('Arial', 'B', 20)
        self.set_text_color(255, 68, 68)  # Rojo #FF4444
        self.cell(0, 10, 'Parrilla 51', ln=False, align='C')
        self.ln(12)
        
        self.set_font('Arial', 'B', 14)
        self.set_text_color(0, 0, 0)
        self.cell(0, 8, self.titulo, ln=True, align='C')
        
        # Fecha de generación
        self.set_font('Arial', 'I', 10)
        self.set_text_color(100, 100, 100)
        self.cell(0, 6, f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', ln=True, align='C')
        
        # Línea separadora
        self.set_draw_color(255, 215, 0)  # Dorado #FFD700
        self.set_line_width(1)
        self.line(10, self.get_y() + 2, 200, self.get_y() + 2)
        self.ln(8)
    
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')

# =====================
# EXPORTAR VENTAS A PDF
# =====================
@reportes_bp.route('/ventas/exportar_pdf')
def exportar_ventas_pdf():
    es_admin, mensaje = verificar_admin()
    if not es_admin:
        flash(mensaje, 'danger')
        return redirect(url_for('auth.login'))
    
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora, p.total, 
               p.estado, p.metodo_pago
        FROM pedidos p
        INNER JOIN usuarios u ON p.cod_usuario = u.id_usuario
        ORDER BY p.fecha DESC
    """)
    pedidos = cursor.fetchall()
    cursor.close()
    conexion.close()

    # Crear PDF personalizado
    pdf = PDF_Parrilla('Reporte de Ventas')
    pdf.add_page()
    
    total_general = 0
    
    # Encabezados de tabla
    pdf.set_font('Arial', 'B', 9)
    pdf.set_fill_color(255, 68, 68)  # Rojo
    pdf.set_text_color(255, 255, 255)
    pdf.cell(20, 8, 'ID', 1, 0, 'C', True)
    pdf.cell(50, 8, 'Cliente', 1, 0, 'C', True)
    pdf.cell(30, 8, 'Fecha', 1, 0, 'C', True)
    pdf.cell(25, 8, 'Hora', 1, 0, 'C', True)
    pdf.cell(30, 8, 'Total', 1, 0, 'C', True)
    pdf.cell(35, 8, 'Estado', 1, 1, 'C', True)
    
    # Datos
    pdf.set_font('Arial', '', 8)
    pdf.set_text_color(0, 0, 0)
    
    for i, pedido in enumerate(pedidos):
        # Alternar colores de fila
        if i % 2 == 0:
            pdf.set_fill_color(240, 240, 240)
        else:
            pdf.set_fill_color(255, 255, 255)
        
        pdf.cell(20, 7, f"#{pedido['id_pedido']}", 1, 0, 'C', True)
        pdf.cell(50, 7, f"{pedido['nombre']} {pedido['apellido']}"[:25], 1, 0, 'L', True)
        pdf.cell(30, 7, str(pedido['fecha']), 1, 0, 'C', True)
        pdf.cell(25, 7, str(pedido['hora'])[:5], 1, 0, 'C', True)
        pdf.cell(30, 7, f"${pedido['total']:,.0f}", 1, 0, 'R', True)
        pdf.cell(35, 7, pedido['estado'][:15], 1, 1, 'C', True)
        
        total_general += pedido['total'] if pedido['total'] else 0
    
    # Total
    pdf.ln(5)
    pdf.set_font('Arial', 'B', 12)
    pdf.set_fill_color(255, 215, 0)  # Dorado
    pdf.set_text_color(0, 0, 0)
    pdf.cell(155, 10, 'TOTAL GENERAL:', 1, 0, 'R', True)
    pdf.cell(35, 10, f'${total_general:,.0f}', 1, 1, 'R', True)

    # Generar PDF en memoria
    salida = BytesIO()
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    salida.write(pdf_bytes)
    salida.seek(0)

    return send_file(
        salida,
        as_attachment=True,
        download_name=f"reporte_ventas_{datetime.now().strftime('%Y%m%d')}.pdf",
        mimetype="application/pdf"
    )

# =====================
# EXPORTAR INVENTARIO A PDF
# =====================
@reportes_bp.route('/inventario/exportar_pdf')
def exportar_inventario_pdf():
    es_admin, mensaje = verificar_admin()
    if not es_admin:
        flash(mensaje, 'danger')
        return redirect(url_for('auth.login'))
    
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute("""
        SELECT p.id_producto, p.nombre, p.cantidad, p.precio, c.nombre_categoria
        FROM productos p
        LEFT JOIN categorias c ON p.cod_categoria = c.id_categoria
        ORDER BY p.cantidad ASC
    """)
    productos = cursor.fetchall()
    cursor.close()
    conexion.close()

    # Crear PDF personalizado
    pdf = PDF_Parrilla('Reporte de Inventario')
    pdf.add_page()
    
    valor_total = 0
    
    # Encabezados de tabla
    pdf.set_font('Arial', 'B', 9)
    pdf.set_fill_color(255, 215, 0)  # Dorado
    pdf.set_text_color(0, 0, 0)
    pdf.cell(15, 8, 'ID', 1, 0, 'C', True)
    pdf.cell(60, 8, 'Producto', 1, 0, 'C', True)
    pdf.cell(40, 8, 'Categoria', 1, 0, 'C', True)
    pdf.cell(25, 8, 'Stock', 1, 0, 'C', True)
    pdf.cell(25, 8, 'Precio', 1, 0, 'C', True)
    pdf.cell(25, 8, 'Valor', 1, 1, 'C', True)
    
    # Datos
    pdf.set_font('Arial', '', 8)
    
    for i, producto in enumerate(productos):
        # Alternar colores de fila
        if i % 2 == 0:
            pdf.set_fill_color(240, 240, 240)
        else:
            pdf.set_fill_color(255, 255, 255)
        
        # Color según stock
        if producto['cantidad'] == 0:
            pdf.set_text_color(220, 53, 69)  # Rojo
        elif producto['cantidad'] < 5:
            pdf.set_text_color(255, 193, 7)  # Amarillo
        else:
            pdf.set_text_color(0, 0, 0)  # Negro
        
        valor_producto = (producto['cantidad'] * producto['precio']) if producto['cantidad'] and producto['precio'] else 0
        
        pdf.cell(15, 7, str(producto['id_producto']), 1, 0, 'C', True)
        pdf.cell(60, 7, producto['nombre'][:30], 1, 0, 'L', True)
        pdf.cell(40, 7, (producto['nombre_categoria'] or 'N/A')[:20], 1, 0, 'C', True)
        pdf.cell(25, 7, str(producto['cantidad']), 1, 0, 'C', True)
        pdf.cell(25, 7, f"${producto['precio']:,.0f}", 1, 0, 'R', True)
        pdf.cell(25, 7, f"${valor_producto:,.0f}", 1, 1, 'R', True)
        
        valor_total += valor_producto
    
    # Total
    pdf.ln(5)
    pdf.set_font('Arial', 'B', 12)
    pdf.set_fill_color(255, 215, 0)  # Dorado
    pdf.set_text_color(0, 0, 0)
    pdf.cell(165, 10, 'VALOR TOTAL DEL INVENTARIO:', 1, 0, 'R', True)
    pdf.cell(25, 10, f'${valor_total:,.0f}', 1, 1, 'R', True)

    # Generar PDF en memoria
    salida = BytesIO()
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    salida.write(pdf_bytes)
    salida.seek(0)

    return send_file(
        salida,
        as_attachment=True,
        download_name=f"reporte_inventario_{datetime.now().strftime('%Y%m%d')}.pdf",
        mimetype="application/pdf"
    )